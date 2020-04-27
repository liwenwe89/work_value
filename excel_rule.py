
import xlrd
import xlwt
from xlutils.copy import copy

import easygui
import datetime
import time

import openpyxl   # openpyxl太卡了
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment,colors
from openpyxl.utils import get_column_letter,column_index_from_string


if __name__ == '__main__':

    #excle 路径
    path=r'行业组项目进展汇总 - 2020.xlsx'


    #打开文件
    excel_book = xlrd.open_workbook(path) 
    wt_excel_bool = copy(excel_book)

    str_sheets =excel_book.sheets()
    hyperlink_name = str_sheets[1].name #第二页名字
    #获取第2个sheet页面，第一个sheet为目录页

    excel_table_1 = excel_book.sheet_by_index(1)


    #获取有效行数
    sheet1_nrows = excel_table_1.nrows
    sheet1_ncols = excel_table_1.ncols
    print(sheet1_nrows)
    print(sheet1_ncols)

    #list_val = [[] for i in range(sheet1_nrows)]

    temp_i = 0 #当前计数
    temp_j = 0 #当前计数

    list_val_single = ["","","","","",""]#项目名称，项目经理，本周进度，下周计划，风险
    list_val = [list_val_single for i in range(100)] #最大二十个项目
    list_link = [str for i in range(100)]
    count = 0
    #每一行都保存为一个列表
    for i in range(0,sheet1_nrows):
        for j in range(0,sheet1_ncols):
            value = excel_table_1.cell(i,j).value
            list_val[count] = ["","","","","",""]
            if (value == r"项目名称"):
                list_val[count][0] = excel_table_1.cell(i,j+1).value #+1为项目名称
                list_link[count] = "B"+str(i+1)
                count = count+1
            if(value == r"汇报人"):
                list_val[count-1][1] = excel_table_1.cell(i,j+1).value
            if(value == r"本周进展"):
                list_val[count-1][2] = excel_table_1.cell(i+1,j).value
            if(value == r"下周计划"):
                list_val[count-1][3] = excel_table_1.cell(i+1,j).value
            if(value == r"处理中"):
                list_val[count-1][4] = list_val[count-1][4]+"问题描述:\n" + str(excel_table_1.cell(i,j-6).value) + "\n 本周进展：\n"+ str(excel_table_1.cell(i,j-3).value)+"\n"
            print("读取数据",i,j,value)


    font_title = Font(name='微软雅黑',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    
# 调整行高

   

    font_cc = Font(name='微软雅黑',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    font_hyperlink = Font(name='微软雅黑',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline="single",
                strike=False,
                color=colors.BLUE)
    
    alignment_excel=Alignment(horizontal='general',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)


    print("\n\n 加载中！！！！\n勿动，可能要好几分钟！！！\n\n")
    workbook_0 = openpyxl.load_workbook(path)
    worksheet_0= workbook_0.worksheets[0]
  


    rows=worksheet_0.max_row #最大行列
    columns=worksheet_0.max_column

    for i in range (1,rows):
        for j in range(1,columns):
            worksheet_0.cell(i,j,"") #清空
    

    print("清除首页",rows,columns,worksheet_0)
    print("重建首页")
    title=['项目名称', '项目经理', '本周进度', '下周计划', '风险']
    for i in range(1,1+len(title)):
        worksheet_0.row_dimensions[1].height = 20
        worksheet_0.cell(1, i,title[i-1])
        worksheet_0.cell(1, i).font = font_title
        worksheet_0.cell(1, i).alignment = alignment_excel


    worksheet_0.column_dimensions[get_column_letter(1)].width = 30
    worksheet_0.column_dimensions[get_column_letter(2)].width = 8
    worksheet_0.column_dimensions[get_column_letter(3)].width = 40
    worksheet_0.column_dimensions[get_column_letter(4)].width = 40
    worksheet_0.column_dimensions[get_column_letter(5)].width = 40
    
   
    worksheet_0.cell(1, 1).alignment = alignment_excel
 #       worksheet_0[1].font =font_title
  #      worksheet_0[1].alignment =alignment_excel
    

    for i in range(2,2+count):   #第二行开始
        worksheet_0.row_dimensions[i].height = 40
 #       worksheet_0[i].font =font_title
 #       worksheet_0[i].alignment = Alignment(horizontal='center', vertical='center')
        for j in range(1,1+len(list_val_single)):    
            worksheet_0.cell(i, j,list_val[i-2][j-1]).font = font_cc
            worksheet_0.cell(i, j,list_val[i-2][j-1]).alignment = alignment_excel
            if (j == 1):
                str_temp = r"#"+hyperlink_name+"!" + str(list_link[i-2])
                worksheet_0.cell(i, j,list_val[i-2][j-1]).hyperlink = str_temp
                worksheet_0.cell(i, j,list_val[i-2][j-1]).font = font_hyperlink
    workbook_0.save(path)

    input("已经完成\n")
    
   
